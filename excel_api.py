# # excel_api.py
# import os, json, re, html, tempfile
# from io import BytesIO
# from typing import List, Dict

# from flask import Flask, request, jsonify, Response
# from openpyxl import load_workbook

# # .msg用
# import extract_msg  # pip install extract-msg

# # .eml用（標準ライブラリ）
# from email import policy
# from email.parser import BytesParser

# app = Flask(__name__)

# # 上限（必要に応じて調整）
# MAX_ROWS = 200
# MAX_COLS = 50
# MAX_NONEMPTY = 2000  # 非空セルの最大数（安全弁）

# def to_str(v) -> str:
#     if v is None:
#         return ""
#     s = str(v)
#     return (
#         s.replace("_x000D_", " ")
#          .replace("\t", " ")
#          .replace("\r\n", " ")
#          .replace("\n", " ")
#          .replace("\r", " ")
#          .strip()
#     )

# def _html_to_text(html_s: str) -> str:
#     if not html_s:
#         return ""
#     s = re.sub(r'(?is)<(script|style).*?>.*?</\1>', '', html_s)
#     s = re.sub(r'(?is)<br\s*/?>', '\n', s)
#     s = re.sub(r'(?is)</p\s*>', '\n', s)
#     s = re.sub(r'(?is)<.*?>', '', s)
#     s = html.unescape(s)
#     return to_str(s)

# def _excel_sparse_from_bytes(xlsx_bytes: bytes,
#                              max_rows=MAX_ROWS, max_cols=MAX_COLS, max_nonempty=MAX_NONEMPTY) -> str:
#     wb = load_workbook(BytesIO(xlsx_bytes), data_only=True, read_only=True)
#     ws = wb.active
#     lines, count = [], 0
#     for row in ws.iter_rows(min_row=1, max_row=max_rows, min_col=1, max_col=max_cols, values_only=False):
#         for cell in row:
#             v = cell.value
#             if v is None:
#                 continue
#             txt = to_str(v)
#             if not txt:
#                 continue
#             lines.append(f"{cell.coordinate}\t{txt}")
#             count += 1
#             if count >= max_nonempty:
#                 lines.append("# ...truncated...")
#                 break
#         if count >= max_nonempty:
#             break
#     return "\n".join(lines)

# def _is_excel_filename(name: str) -> bool:
#     n = (name or "").lower()
#     return n.endswith(".xlsx") or n.endswith(".xlsm")

# def _is_excel_mime(mime: str) -> bool:
#     return (mime or "").lower().startswith("application/vnd.openxmlformats-officedocument.spreadsheetml")

# @app.route("/", methods=["GET"])
# def health():
#     return jsonify({
#         "ok": True,
#         "message": "excel-api (xlsx sparse + mail .msg/.eml)",
#         "endpoint": ["/extract", "/extract_mail"]
#     })

# @app.route("/extract", methods=["POST"])
# def extract():
#     """
#     既存：.xlsx直読み → 非空セルのみ 'A1<TAB>値' のTSVを返す
#     multipart/form-data:
#       - file: (必須) .xlsx
#       - bom: "true"/"false" 既定 true
#       - inline: "true"/"false" 既定 true
#       - sheet: オプション（名前 / 0始まり / 1始まり）
#     """
#     f = request.files.get("file")
#     if not f:
#         return jsonify({"error": "file is required (multipart/form-data)"}), 400

#     bom_on = (request.form.get("bom", "true").lower() != "false")
#     inline_on = (request.form.get("inline", "true").lower() != "false")
#     sheet_req = request.form.get("sheet")

#     try:
#         wb = load_workbook(f, data_only=True, read_only=True)
#     except Exception as e:
#         return jsonify({"error": f"failed to read workbook: {e}"}), 400

#     ws = None
#     if sheet_req:
#         try:
#             idx = int(sheet_req)
#             names = wb.sheetnames
#             if 0 <= idx < len(names):
#                 ws = wb[names[idx]]
#             elif 1 <= idx <= len(names):
#                 ws = wb[names[idx - 1]]
#         except ValueError:
#             if sheet_req in wb.sheetnames:
#                 ws = wb[sheet_req]
#     ws = ws or wb.active

#     lines, count = [], 0
#     for row in ws.iter_rows(min_row=1, max_row=MAX_ROWS, min_col=1, max_col=MAX_COLS, values_only=False):
#         for cell in row:
#             v = cell.value
#             if v is None:
#                 continue
#             txt = to_str(v)
#             if not txt:
#                 continue
#             lines.append(f"{cell.coordinate}\t{txt}")
#             count += 1
#             if count >= MAX_NONEMPTY:
#                 lines.append("# ...truncated...")
#                 break
#         if count >= MAX_NONEMPTY:
#             break

#     payload = "\n".join(lines)
#     if bom_on:
#         payload = "\ufeff" + payload

#     headers = {}
#     if not inline_on:
#         headers["Content-Disposition"] = 'attachment; filename="extract.tsv"'
#     return Response(payload, mimetype="text/plain; charset=utf-8", headers=headers)

# @app.route("/extract_mail", methods=["POST"])
# def extract_mail():
#     """
#     .msg/.eml を自動判別して、
#       - body_text（本文プレーンテキスト）
#       - excel_attachments: [{filename, cells}]
#     を JSON で返す。

#     multipart/form-data:
#       - file: (必須) .msg or .eml
#     """
#     up = request.files.get("file")
#     if not up:
#         return jsonify({"error": "file is required (multipart/form-data)"}), 400

#     filename = (up.filename or "").lower()
#     data = up.read()

#     try:
#         if filename.endswith(".msg") or _looks_like_msg(data):
#             payload = _handle_msg_bytes(data)
#         elif filename.endswith(".eml") or _looks_like_eml(data):
#             payload = _handle_eml_bytes(data)
#         else:
#             # 拡張子/シグネチャ不明 → .emlとして最終トライ
#             try:
#                 payload = _handle_eml_bytes(data)
#             except Exception:
#                 # だめなら .msg も試す
#                 try:
#                     payload = _handle_msg_bytes(data)
#                 except Exception as e:
#                     return jsonify({"error": f"unsupported or unreadable mail file: {e}"}), 400
#         return Response(json.dumps(payload, ensure_ascii=False),
#                         mimetype="application/json; charset=utf-8")
#     except Exception as e:
#         return jsonify({"error": f"failed to process mail: {e}"}), 400

# def _looks_like_msg(b: bytes) -> bool:
#     # .msg は OLE (D0 CF 11 E0 A1 B1 1A E1) マジックを持つことが多い
#     return len(b) >= 8 and b[:8] == b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"

# def _looks_like_eml(b: bytes) -> bool:
#     # ざっくり：ASCIIテキスト + ヘッダ行が並ぶ感じ（超簡易判定）
#     head = b[:512].decode("utf-8", errors="ignore")
#     return ("From:" in head or "Subject:" in head) and "\n\n" in head

# def _handle_msg_bytes(b: bytes) -> Dict:
#     # extract-msg はパス読みが安定なので一時ファイルへ
#     with tempfile.NamedTemporaryFile(delete=True, suffix=".msg") as tmp:
#         tmp.write(b)
#         tmp.flush()
#         msg = extract_msg.Message(tmp.name)

#     raw_text = to_str(getattr(msg, "body", "") or "")
#     raw_html = getattr(msg, "bodyHTML", "") or ""
#     body_text = raw_text or _html_to_text(raw_html)

#     excel_results: List[Dict] = []
#     for att in msg.attachments:
#         name = getattr(att, "longFilename", "") or getattr(att, "shortFilename", "") or "attachment"
#         data = getattr(att, "data", None)
#         if not data:
#             continue
#         if _is_excel_filename(name):
#             try:
#                 cells_text = _excel_sparse_from_bytes(data)
#             except Exception as e:
#                 cells_text = f"# ERROR: excel parse failed: {e}"
#             excel_results.append({"filename": name, "cells": cells_text})
#     return {"ok": True, "format": "msg", "body_text": body_text, "excel_attachments": excel_results}

# def _handle_eml_bytes(b: bytes) -> Dict:
#     msg = BytesParser(policy=policy.default).parsebytes(b)

#     # 本文：text/plain を優先、無ければ text/html をテキスト化
#     body_text = ""
#     if msg.is_multipart():
#         # text/plain を探す
#         for part in msg.walk():
#             if part.get_content_type() == "text/plain" and part.get_content_disposition() in (None, "inline"):
#                 body_text = to_str(part.get_content())
#                 if body_text:
#                     break
#         if not body_text:
#             for part in msg.walk():
#                 if part.get_content_type() == "text/html" and part.get_content_disposition() in (None, "inline"):
#                     body_text = _html_to_text(part.get_content())
#                     if body_text:
#                         break
#     else:
#         ctype = msg.get_content_type()
#         if ctype == "text/plain":
#             body_text = to_str(msg.get_content())
#         elif ctype == "text/html":
#             body_text = _html_to_text(msg.get_content())

#     # 添付Excel
#     excel_results: List[Dict] = []
#     for part in msg.walk():
#         fname = part.get_filename()  # RFC2231/2047 を policy=default がよしなに復号
#         cdisp = part.get_content_disposition()
#         ctype = part.get_content_type()
#         if cdisp == "attachment" or fname:
#             if _is_excel_filename(fname) or _is_excel_mime(ctype):
#                 data = part.get_payload(decode=True) or b""
#                 if not data:
#                     continue
#                 try:
#                     cells_text = _excel_sparse_from_bytes(data)
#                 except Exception as e:
#                     cells_text = f"# ERROR: excel parse failed: {e}"
#                 excel_results.append({"filename": fname or "attachment.xlsx", "cells": cells_text})

#     return {"ok": True, "format": "eml", "body_text": body_text, "excel_attachments": excel_results}

# if __name__ == "__main__":
#     port = int(os.environ.get("PORT", "10000"))
#     app.run(host="0.0.0.0", port=port)





# excel_api.py
import os, json, re, html, tempfile
from io import BytesIO
from typing import List, Dict

from flask import Flask, request, jsonify, Response
from openpyxl import load_workbook
import xlrd  # ← .xls対応

# .msg用
import extract_msg  # pip install extract-msg

# .eml用（標準ライブラリ）
from email import policy
from email.parser import BytesParser

app = Flask(__name__)

# 上限（必要に応じて調整）
MAX_ROWS = 200
MAX_COLS = 50
MAX_NONEMPTY = 2000  # 非空セルの最大数（安全弁）

def to_str(v) -> str:
    if v is None:
        return ""
    s = str(v)
    return (
        s.replace("_x000D_", " ")
         .replace("\t", " ")
         .replace("\r\n", " ")
         .replace("\n", " ")
         .replace("\r", " ")
         .strip()
    )

def _html_to_text(html_s: str) -> str:
    if not html_s:
        return ""
    s = re.sub(r'(?is)<(script|style).*?>.*?</\1>', '', html_s)
    s = re.sub(r'(?is)<br\s*/?>', '\n', s)
    s = re.sub(r'(?is)</p\s*>', '\n', s)
    s = re.sub(r'(?is)<.*?>', '', s)
    s = html.unescape(s)
    return to_str(s)

# ========= Excel 抽出ロジック =========

def _num_to_col(n: int) -> str:
    # 1->A, 26->Z, 27->AA...
    s = []
    while n > 0:
        n, rem = divmod(n - 1, 26)
        s.append(chr(65 + rem))
    return "".join(reversed(s))

def _excel_sparse_from_xlsx_bytes(xlsx_bytes: bytes,
                                  sheet_req: str | None = None,
                                  max_rows=MAX_ROWS, max_cols=MAX_COLS, max_nonempty=MAX_NONEMPTY) -> str:
    wb = load_workbook(BytesIO(xlsx_bytes), data_only=True, read_only=True)
    ws = None
    if sheet_req:
        try:
            idx = int(sheet_req)
            names = wb.sheetnames
            if 0 <= idx < len(names):
                ws = wb[names[idx]]
            elif 1 <= idx <= len(names):
                ws = wb[names[idx - 1]]
        except ValueError:
            if sheet_req in wb.sheetnames:
                ws = wb[sheet_req]
    ws = ws or wb.active

    lines, count = [], 0
    for row in ws.iter_rows(min_row=1, max_row=max_rows, min_col=1, max_col=max_cols, values_only=False):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            txt = to_str(v)
            if not txt:
                continue
            lines.append(f"{cell.coordinate}\t{txt}")
            count += 1
            if count >= max_nonempty:
                lines.append("# ...truncated...")
                break
        if count >= max_nonempty:
            break
    return "\n".join(lines)

def _excel_sparse_from_xls_bytes(xls_bytes: bytes,
                                 max_rows=MAX_ROWS, max_cols=MAX_COLS, max_nonempty=MAX_NONEMPTY) -> str:
    with tempfile.NamedTemporaryFile(delete=True, suffix=".xls") as tmp:
        tmp.write(xls_bytes)
        tmp.flush()
        book = xlrd.open_workbook(tmp.name)
        sheet = book.sheet_by_index(0)  # 先頭シートのみ

        lines, count = [], 0
        max_r = min(sheet.nrows, max_rows)
        max_c = min(sheet.ncols, max_cols)
        for r in range(max_r):
            for c in range(max_c):
                v = sheet.cell_value(r, c)
                txt = to_str(v)
                if not txt:
                    continue
                coord = f"{_num_to_col(c+1)}{r+1}"
                lines.append(f"{coord}\t{txt}")
                count += 1
                if count >= max_nonempty:
                    lines.append("# ...truncated...")
                    break
            if count >= max_nonempty:
                break
    return "\n".join(lines)

def _excel_sparse_from_bytes(data: bytes,
                             filename: str | None = None,
                             sheet_req: str | None = None,
                             max_rows=MAX_ROWS, max_cols=MAX_COLS, max_nonempty=MAX_NONEMPTY) -> str:
    name = (filename or "").lower()
    if name.endswith(".xls"):
        return _excel_sparse_from_xls_bytes(data, max_rows, max_cols, max_nonempty)
    return _excel_sparse_from_xlsx_bytes(data, sheet_req, max_rows, max_cols, max_nonempty)

def _is_excel_filename(name: str) -> bool:
    n = (name or "").lower()
    return n.endswith((".xlsx", ".xlsm", ".xls"))

def _is_excel_mime(mime: str) -> bool:
    m = (mime or "").lower()
    return (
        m.startswith("application/vnd.openxmlformats-officedocument.spreadsheetml")
        or m == "application/vnd.ms-excel"
    )

# ========= Flaskエンドポイント =========

@app.route("/", methods=["GET"])
def health():
    return jsonify({
        "ok": True,
        "message": "excel-api (xlsx/xls sparse + mail .msg/.eml)",
        "endpoint": ["/extract", "/extract_mail"]
    })

@app.route("/extract", methods=["POST"])
def extract():
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "file is required (multipart/form-data)"}), 400

    bom_on = (request.form.get("bom", "true").lower() != "false")
    inline_on = (request.form.get("inline", "true").lower() != "false")
    sheet_req = request.form.get("sheet")

    data = f.read()
    if not data:
        return jsonify({"error": "empty file"}), 400

    try:
        payload = _excel_sparse_from_bytes(data, filename=f.filename, sheet_req=sheet_req)
    except Exception as e:
        return jsonify({"error": f"failed to read workbook: {e}"}), 400

    if bom_on:
        payload = "\ufeff" + payload

    headers = {}
    if not inline_on:
        headers["Content-Disposition"] = 'attachment; filename="extract.tsv"'
    return Response(payload, mimetype="text/plain; charset=utf-8", headers=headers)

@app.route("/extract_mail", methods=["POST"])
def extract_mail():
    up = request.files.get("file")
    if not up:
        return jsonify({"error": "file is required (multipart/form-data)"}), 400

    filename = (up.filename or "").lower()
    data = up.read()

    try:
        if filename.endswith(".msg") or _looks_like_msg(data):
            payload = _handle_msg_bytes(data)
        elif filename.endswith(".eml") or _looks_like_eml(data):
            payload = _handle_eml_bytes(data)
        else:
            try:
                payload = _handle_eml_bytes(data)
            except Exception:
                try:
                    payload = _handle_msg_bytes(data)
                except Exception as e:
                    return jsonify({"error": f"unsupported or unreadable mail file: {e}"}), 400
        return Response(json.dumps(payload, ensure_ascii=False),
                        mimetype="application/json; charset=utf-8")
    except Exception as e:
        return jsonify({"error": f"failed to process mail: {e}"}), 400

# ========= メール処理 =========

def _looks_like_msg(b: bytes) -> bool:
    return len(b) >= 8 and b[:8] == b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"

def _looks_like_eml(b: bytes) -> bool:
    head = b[:512].decode("utf-8", errors="ignore")
    return ("From:" in head or "Subject:" in head) and "\n\n" in head

def _handle_msg_bytes(b: bytes) -> Dict:
    with tempfile.NamedTemporaryFile(delete=True, suffix=".msg") as tmp:
        tmp.write(b)
        tmp.flush()
        msg = extract_msg.Message(tmp.name)

    raw_text = to_str(getattr(msg, "body", "") or "")
    raw_html = getattr(msg, "bodyHTML", "") or ""
    body_text = raw_text or _html_to_text(raw_html)

    excel_results: List[Dict] = []
    for att in msg.attachments:
        name = getattr(att, "longFilename", "") or getattr(att, "shortFilename", "") or "attachment"
        data = getattr(att, "data", None)
        if not data:
            continue
        if _is_excel_filename(name):
            try:
                cells_text = _excel_sparse_from_bytes(data, filename=name)
            except Exception as e:
                cells_text = f"# ERROR: excel parse failed: {e}"
            excel_results.append({"filename": name, "cells": cells_text})
    return {"ok": True, "format": "msg", "body_text": body_text, "excel_attachments": excel_results}

def _handle_eml_bytes(b: bytes) -> Dict:
    msg = BytesParser(policy=policy.default).parsebytes(b)

    body_text = ""
    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_type() == "text/plain" and part.get_content_disposition() in (None, "inline"):
                body_text = to_str(part.get_content())
                if body_text:
                    break
        if not body_text:
            for part in msg.walk():
                if part.get_content_type() == "text/html" and part.get_content_disposition() in (None, "inline"):
                    body_text = _html_to_text(part.get_content())
                    if body_text:
                        break
    else:
        ctype = msg.get_content_type()
        if ctype == "text/plain":
            body_text = to_str(msg.get_content())
        elif ctype == "text/html":
            body_text = _html_to_text(msg.get_content())

    excel_results: List[Dict] = []
    for part in msg.walk():
        fname = part.get_filename()
        cdisp = part.get_content_disposition()
        ctype = part.get_content_type()
        if cdisp == "attachment" or fname:
            if _is_excel_filename(fname) or _is_excel_mime(ctype):
                data = part.get_payload(decode=True) or b""
                if not data:
                    continue
                try:
                    cells_text = _excel_sparse_from_bytes(data, filename=fname or "")
                except Exception as e:
                    cells_text = f"# ERROR: excel parse failed: {e}"
                excel_results.append({"filename": fname or "attachment.xlsx", "cells": cells_text})

    return {"ok": True, "format": "eml", "body_text": body_text, "excel_attachments": excel_results}

# ========= main =========

if __name__ == "__main__":
    port = int(os.environ.get("PORT", "10000"))
    app.run(host="0.0.0.0", port=port)
